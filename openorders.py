import openpyxl

# Define file location
path = "openorders.xlsx"

# Create a workbook object and open the excel file
wb_obj = openpyxl.load_workbook(path)

# Create a worksheet object and assign the active sheet to it
sheet_obj = wb_obj.active

# Get the total number of rows
row = sheet_obj.max_row

# Iterate through all rows to identify any rows where status = 'A'
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2).value
    if cell_obj == "A":
        order = sheet_obj.cell(row=i, column=1).value
        sku = sheet_obj.cell(row=i, column=3).value
        total = sheet_obj.cell(row=i, column=4).value
        picked = sheet_obj.cell(row=i, column=5).value
        inv = sheet_obj.cell(row=i, column=6).value

        # Determine if order can be fulfilled based on inventory
        remaining = total - picked
        if remaining <= inv:
            print("Order Number: " + order)
            print("SKU: " + str(int(sku)))
            print("Order Total: " + str(int(total)))
            print("Amount Picked: " + str(int(picked)))
            print("Amount in Inventory: " + str(int(inv)))
            print("Amount to pick: " + str(int(remaining)))
            print("")
